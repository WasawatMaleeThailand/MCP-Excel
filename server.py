# server.py
import os
import uuid
from typing import Dict, Any, List
from fastapi import FastAPI, HTTPException, Request
import openpyxl
from openpyxl.utils import column_index_from_string

app = FastAPI(title="Excel MCP Server on Codespace")

# --- อัปเดต Tool ที่เซิร์ฟเวอร์นี้มีความสามารถ ---
TOOLS = [
    {
        "name": "write_data_to_excel",
        "description": "Write data to an Excel worksheet.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "filepath": {"type": "string", "description": "Path to Excel file"},
                "sheet_name": {"type": "string", "description": "Name of worksheet"},
                "data": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "List of lists representing rows of data"
                },
                "start_cell": {"type": "string", "description": "Cell to start writing, e.g., 'A1'"}
            },
            "required": ["filepath", "sheet_name", "data"]
        }
    }
]

def excel_writer(filepath: str, sheet_name: str, data: List[List[Any]], start_cell: str = "A1"):
    """ฟังก์ชันสำหรับเขียนข้อมูลลงไฟล์ Excel"""
    try:
        # ตรวจสอบว่าไฟล์มีอยู่หรือไม่
        if os.path.exists(filepath):
            workbook = openpyxl.load_workbook(filepath)
        else:
            workbook = openpyxl.Workbook()
            # ลบชีตเริ่มต้นถ้ามี
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        # เลือกชีตหรือสร้างใหม่ถ้ายังไม่มี
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        # แปลง start_cell (เช่น "A1") เป็น row และ column index
        start_col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row_num = ''.join(filter(str.isdigit, start_cell))
        start_col = column_index_from_string(start_col_letter)
        start_row = int(start_row_num)

        # วนลูปเพื่อเขียนข้อมูลลงในแต่ละ cell
        for r_idx, row_data in enumerate(data, start=start_row):
            for c_idx, cell_value in enumerate(row_data, start=start_col):
                sheet.cell(row=r_idx, column=c_idx, value=cell_value)

        # บันทึกไฟล์
        workbook.save(filepath)
        return {"status": "success", "message": f"Data written to {filepath} in sheet '{sheet_name}'"}

    except Exception as e:
        # หากเกิดข้อผิดพลาด ให้โยน Exception ออกไป
        raise e

def frame_reply(in_reply_to: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    """ฟังก์ชันช่วยสร้าง Frame ตอบกลับ"""
    out = {"inReplyTo": in_reply_to}
    out.update(payload)
    return out

# --- Endpoint หลักของเซิร์ฟเวอร์ ---
@app.post("/mcp")
async def mcp_endpoint(req: Request):
    try:
        frames = await req.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    if not isinstance(frames, list):
        raise HTTPException(400, "Payload must be a JSON array of frames")

    replies: List[Dict[str, Any]] = []

    for f in frames:
        ftype = f.get("type")
        fid = f.get("id") or str(uuid.uuid4())

        if ftype == "initialize":
            replies.append(frame_reply(fid, {"type": "initialized"}))
        elif ftype == "listTools":
            replies.append(frame_reply(fid, {"type": "tools", "tools": TOOLS}))
        elif ftype == "callTool":
            # --- อัปเดต Logic การเรียกใช้ Tool ---
            tool_name = f.get("name")
            if tool_name == "write_data_to_excel":
                try:
                    args = f.get("arguments", {})
                    result = excel_writer(
                        filepath=args.get("filepath"),
                        sheet_name=args.get("sheet_name"),
                        data=args.get("data"),
                        start_cell=args.get("start_cell", "A1") # ใช้ "A1" เป็นค่า default
                    )
                    replies.append(frame_reply(fid, {"type": "toolResult", "result": result}))
                except Exception as e:
                    replies.append(frame_reply(fid, {"type": "error", "error": str(e)}))
            else:
                replies.append(frame_reply(fid, {"type": "error", "error": f"Unknown tool: {tool_name}"}))
        else:
            replies.append(frame_reply(fid, {"type": "error", "error": f"Unknown frame type: {ftype}"}))

    return replies
